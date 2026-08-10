#!/usr/bin/env python3
"""Extract playing owners + retained players (with category & pre-auction cost)
from the organizers' "Team Owners" sheet. Category is encoded as the cell FILL
COLOUR (SheetJS can't read these, so we use openpyxl):
    yellow FFFFFF00 = Legend, theme4 = Category A, theme9 = Category B, no fill = non-playing.
Legend is treated as B for cost. Non-playing owners (no fill / the dedicated
"Non-Playing Owner" column) are skipped entirely — they never touch the purse.
Cost tiers:  owner  A ₹15,000 · B/Legend ₹6,000
             retained A ₹20,000 · B/Legend ₹10,000
Writes scripts/sccl-roster.json.  Run: python3 scripts/extract-roster.py <xlsx>
"""
import openpyxl, json, sys, os

XLSX = sys.argv[1] if len(sys.argv) > 1 else \
    "/Users/nikhildhingra/Downloads/For_owners_final_auction_list_7th_Aug.xlsx"
OUT = os.path.join(os.path.dirname(__file__), "sccl-roster.json")

wb = openpyxl.load_workbook(XLSX)  # keep styles (no data_only)
ws = wb["Team Owners"]

def category(cell):
    f = cell.fill
    if f is None or f.fgColor is None:
        return None
    if f.fgColor.type == "rgb" and f.fgColor.rgb == "FFFFFF00":
        return "Legend"
    if f.fgColor.type == "theme":
        return {4: "A", 9: "B"}.get(f.fgColor.theme)
    return None

def cost(kind, cat):
    is_a = cat == "A"
    return (15000 if is_a else 6000) if kind == "owner" else (20000 if is_a else 10000)

OWNER_COLS = [4, 5, 6]      # Owner 1..3   (col 7 = Non-Playing Owner: skipped)
RET_COLS = [8, 9, 10, 11]   # Retained 1..4

roster = []
for r in range(2, 26):
    team = str(ws.cell(row=r, column=2).value or "").strip()
    if not team:
        continue
    for c in OWNER_COLS:
        name = str(ws.cell(row=r, column=c).value or "").strip()
        cat = category(ws.cell(row=r, column=c))
        if not name or cat is None:   # no fill = non-playing owner
            continue
        roster.append({"team": team, "name": name, "kind": "owner",
                       "category": cat, "cost": cost("owner", cat)})
    for c in RET_COLS:
        name = str(ws.cell(row=r, column=c).value or "").strip()
        if not name:
            continue
        name = name.split(" - replaced")[0].strip()
        cat = category(ws.cell(row=r, column=c)) or "B"
        roster.append({"team": team, "name": name, "kind": "retained",
                       "category": cat, "cost": cost("retained", cat)})

json.dump(roster, open(OUT, "w"), indent=1, ensure_ascii=False)
print(f"wrote {len(roster)} entries -> {OUT}")
